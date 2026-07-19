#!/usr/bin/env python3
"""Build the Excel tracker from the latest jobs_scored CSV.
Sheets: Top Picks (tier 5 + why) / All Scored / Summary (static snapshot values —
no LibreOffice recalc dependency) / Coverage (per-source, from run_report.json).
Status/Notes columns are preserved across rebuilds (matched by apply-link URL)."""
import csv, glob, json, os
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
src_csv = sorted(glob.glob(os.path.join(HERE, "results", "jobs_scored_*.csv")))[-1]
stamp = src_csv.split("_")[-1].replace(".csv", "")
with open(src_csv) as _f:
    rows = list(csv.DictReader(_f))
rows.sort(key=lambda r: (-int(r["tier"]), r["company"], r["title"]))
report = {}
rp = os.path.join(HERE, "results", "run_report.json")
if os.path.exists(rp):
    with open(rp) as _f:
        report = json.load(_f)


def _safe(v):
    """Neutralize spreadsheet formula injection: scraped titles/companies/descriptions flow
    into cells, and Excel/openpyxl treats a leading =,+,-,@ (or a control char) as a formula.
    Prefix those with an apostrophe so the value renders as literal text."""
    if isinstance(v, str) and v and (v[0] in "=+-@\t\r" or v[:1] == "\x00"):
        return "'" + v
    return v

def _safe_url(u):
    """Only allow http(s) apply links to become live hyperlinks (blocks file:/javascript: etc.)."""
    return u if isinstance(u, str) and u.lower().startswith(("http://", "https://")) else None

def _col_index(ws, header):
    """0-based index of a column by its header cell text (layout-robust)."""
    for j, c in enumerate(ws[1]):
        if (c.value or "").strip().lower() == header.lower():
            return j
    return None

def load_prev_status():
    """Status/Notes the user typed into earlier trackers, keyed by apply URL.
    Columns are located by header name so layout changes don't break carryover.
    Oldest -> newest so the most recent tracker wins."""
    st = {}
    for path in sorted(glob.glob(os.path.join(HERE, "results", "Job_Hunt_Tracker_*.xlsx"))):
        try:
            pwb = load_workbook(path)
        except Exception:
            continue
        if "All Scored" in pwb.sheetnames:
            ws = pwb["All Scored"]
            li, si, ni = _col_index(ws, "Apply Link"), _col_index(ws, "Status"), _col_index(ws, "Notes")
            for row in ws.iter_rows(min_row=2):
                link = row[li].hyperlink.target if li is not None and len(row) > li and row[li].hyperlink else None
                status = row[si].value if si is not None and len(row) > si else None
                notes = row[ni].value if ni is not None and len(row) > ni else None
                if link and (status or notes):
                    st[link] = {"status": status or "", "notes": notes or ""}
        if "Top Picks" in pwb.sheetnames:
            ws = pwb["Top Picks"]
            li, si = _col_index(ws, "Apply Link"), _col_index(ws, "Status")
            for row in ws.iter_rows(min_row=2):
                link = row[li].hyperlink.target if li is not None and len(row) > li and row[li].hyperlink else None
                status = row[si].value if si is not None and len(row) > si else None
                if link and status:
                    st.setdefault(link, {"status": "", "notes": ""})
                    if not st[link]["status"]:
                        st[link]["status"] = status
    return st


prev = load_prev_status()

ARIAL = Font(name="Arial")
hfill = PatternFill("solid", fgColor="1F3864")
t5fill = PatternFill("solid", fgColor="C6EFCE")
t4fill = PatternFill("solid", fgColor="E2EFDA")
# odds axis (orthogonal to fit tier): green=Likely, amber=Target, red=Reach
ODDS_FILL = {"Likely": PatternFill("solid", fgColor="92D050"),
             "Target": PatternFill("solid", fgColor="FFE699"),
             "Reach":  PatternFill("solid", fgColor="F4B7B7")}
ODDS_RANK = {"Likely": 0, "Target": 1, "Reach": 2}

def odds_sort_key(r):
    # most-winnable first; break ties by numeric odds_score then company
    try: sc = int(r.get("odds_score") or 0)
    except (TypeError, ValueError): sc = 0
    return (ODDS_RANK.get(r.get("odds", "Target"), 1), -sc, r["company"], r["title"])

def style_header(ws):
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True, color="FFFFFF"); c.fill = hfill

def sal(r):
    if r["salary"]: return r["salary"]
    if r["salary_min"]:
        return f"${int(float(r['salary_min'])):,}-${int(float(r['salary_max'])):,}"
    return ""

wb = Workbook()
status_dv = DataValidation(type="list", formula1='"Interested,Applied,Phone screen,Interview,Offer,Rejected,Passed"', allow_blank=True)

# --- Top Picks (sorted most-winnable first; Odds is the hireability axis) ---
tp = wb.active; tp.title = "Top Picks"
tp.append(["Odds", "New", "Title", "Company", "Location", "Salary", "Why it fits", "Why (odds)", "Flags", "Apply Link", "Status"])
style_header(tp); tp.add_data_validation(status_dv)
LINK_COL, STATUS_COL = 10, 11
for r in sorted([r for r in rows if r["tier"] == "5"], key=odds_sort_key):
    tp.append([r.get("odds", ""), r.get("new", ""), _safe(r["title"]), _safe(r["company"]), _safe(r["location"]),
               _safe(sal(r)), _safe(r["why"]), _safe(r.get("odds_why", "")), _safe(r["flags"]), "Apply",
               _safe(prev.get(r["url"], {}).get("status", ""))])
    i = tp.max_row
    link = tp.cell(row=i, column=LINK_COL); hl = _safe_url(r["url"])
    if hl:
        link.hyperlink = hl
        link.font = Font(name="Arial", color="0563C1", underline="single")
    else:
        link.value = "Apply (link withheld)"; link.font = ARIAL
    status_dv.add(tp.cell(row=i, column=STATUS_COL))
    for col in range(1, 12):
        if col != LINK_COL:
            tp.cell(row=i, column=col).font = ARIAL
            tp.cell(row=i, column=col).alignment = Alignment(wrap_text=(col in (7, 8)), vertical="top")
        tp.cell(row=i, column=col).fill = t5fill
    tp.cell(row=i, column=1).fill = ODDS_FILL.get(r.get("odds", "Target"), t5fill)  # odds cell overrides
for i, w in enumerate([9, 6, 44, 17, 24, 20, 50, 34, 15, 10, 13], 1):
    tp.column_dimensions[get_column_letter(i)].width = w
tp.freeze_panes = "A2"

# --- All Scored (Tier = fit, Odds = hireability; filter/sort on either) ---
ws = wb.create_sheet("All Scored")
hdr = ["Tier", "Odds", "New", "Title", "Company", "Location", "Salary", "Posted", "First Seen", "Remote",
       "Source", "Flags", "Why (auto)", "Apply Link", "Req ID", "Status", "Notes", "Why (odds)"]
ws.append(hdr); style_header(ws); ws.add_data_validation(status_dv)
LINK_C, STATUS_C = 14, 16
for r in rows:
    ws.append([int(r["tier"]), r.get("odds", ""), r.get("new", ""), _safe(r["title"]), _safe(r["company"]),
               _safe(r["location"]), _safe(sal(r)), _safe(r["posted"]), r.get("first_seen", ""),
               "Yes" if r["remote"] == "True" else "", _safe(r["source"]), _safe(r["flags"]), _safe(r["why"]),
               "Apply", _safe(r["req_id"]),
               _safe(prev.get(r["url"], {}).get("status", "")), _safe(prev.get(r["url"], {}).get("notes", "")),
               _safe(r.get("odds_why", ""))])
    i = ws.max_row
    link = ws.cell(row=i, column=LINK_C); hl = _safe_url(r["url"])
    if hl:
        link.hyperlink = hl
        link.font = Font(name="Arial", color="0563C1", underline="single")
    else:
        link.value = "Apply (link withheld)"; link.font = ARIAL
    status_dv.add(ws.cell(row=i, column=STATUS_C))
    fill = t5fill if r["tier"] == "5" else (t4fill if r["tier"] == "4" else None)
    for col in range(1, 19):
        if col != LINK_C: ws.cell(row=i, column=col).font = ARIAL
        if fill: ws.cell(row=i, column=col).fill = fill
    ws.cell(row=i, column=2).fill = ODDS_FILL.get(r.get("odds", "Target"), fill or PatternFill())  # odds cell
for i, w in enumerate([6, 8, 6, 46, 20, 26, 18, 11, 11, 8, 14, 20, 34, 10, 12, 13, 22, 30], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:R{ws.max_row}"

# --- Summary (static snapshot; no recalc dependency) ---
s = wb.create_sheet("Summary")
tiers = Counter(int(r["tier"]) for r in rows)
s["A1"] = f"Rubric-scored sweep — {stamp} (snapshot values)"; s["A1"].font = Font(name="Arial", bold=True, size=14)
stats = [("Total scored jobs", len(rows)),
         ("Tier 5 — Apply today (judgment-confirmed)", tiers[5]),
         ("Tier 4 — Strong", tiers[4]),
         ("Tier 3 — Solid / desc unavailable", tiers[3]),
         ("Tier 2 — Adjacent", tiers[2]),
         ("Tier 1 — Volume", tiers[1]),
         ("New this run", sum(1 for r in rows if r.get("new"))),
         ("Salary recovered from descriptions", sum(1 for r in rows if "salary-from-desc" in r["flags"])),
         ("Staffing/W2-flagged (kept per rubric)", sum(1 for r in rows if "Staffing" in r["flags"])),
         ("Descriptions read", sum(1 for r in rows if r.get("desc_snippet")))]
for i, (label, val) in enumerate(stats, start=3):
    s[f"A{i}"] = label; s[f"B{i}"] = val
    s[f"A{i}"].font = ARIAL; s[f"B{i}"].font = Font(name="Arial", bold=True)
s[f"A{len(stats)+4}"] = "Tiers per RUBRIC.md. Tier 5 requires a read description, level fit, domain overlap, and a written justification."
s[f"A{len(stats)+4}"].font = Font(name="Arial", italic=True, size=9)
s.column_dimensions["A"].width = 46; s.column_dimensions["B"].width = 12

# --- Coverage ---
cv = wb.create_sheet("Coverage")
cv.append(["Source", "Kept", "With Description", "Desc %"])
style_header(cv)
for src, d in sorted((report.get("by_source") or {}).items()):
    pct = round(100 * d["with_desc"] / d["kept"]) if d["kept"] else 0
    cv.append([src, d["kept"], d["with_desc"], f"{pct}%"])
    for c in cv[cv.max_row]: c.font = ARIAL
cv.append([]); cv.append(["Blocked (excluded with reason)", "", "", ""])
cv[cv.max_row][0].font = Font(name="Arial", bold=True)
for reason, n in sorted((report.get("blocked") or {}).items(), key=lambda x: -x[1]):
    cv.append([reason, n, "", ""])
    for c in cv[cv.max_row]: c.font = ARIAL
for i, w in enumerate([44, 10, 18, 10], 1):
    cv.column_dimensions[get_column_letter(i)].width = w

out = os.path.join(HERE, "results", f"Job_Hunt_Tracker_{stamp}.xlsx")
try:
    wb.save(out)
except PermissionError:
    # workbook open in Excel/LibreOffice — don't die mid-sweep, save alongside
    out = out.replace(".xlsx", "_new.xlsx")
    wb.save(out)
    print("WARNING: tracker was open in another program; saved as _new.xlsx instead")
print(f"saved {out}: {sum(1 for r in rows if r['tier']=='5')} top picks, {len(rows)} scored, "
      f"{sum(1 for r in rows if prev.get(r['url'], {}).get('status'))} statuses carried over")
